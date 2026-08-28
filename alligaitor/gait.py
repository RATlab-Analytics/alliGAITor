"""Gait metrics computed from triangulated 3D paw trajectories.

Consumes the ``pose_3d.csv`` written by :func:`alligaitor.pipeline.run_session`
(frame, time_s, node, x, y, z, reprojection_error_px) and derives, per rat
crossing:

* total time to cross the platform,
* average speed,
* average stride length per paw (liftoff position to the following
  touchdown position),
* average step length per paw (forward distance from a paw's touchdown to
  the contralateral paw's most recent prior touchdown), and
* average ground contact time per paw (touchdown to liftoff duration).

Coordinates are taken directly from the 3D reconstruction, which is
already metric, so no separate pixel-to-length ratio is needed.
Ground-contact detection is speed-only (see
:class:`alligaitor.config.GaitConfig`); a height-above-platform check was
tried and dropped since it needs each frame's actual depth to be accurate.

:func:`alligaitor.pipeline.run_group` runs the full 2D/3D pipeline for
every session in a group, computes each session's :class:`TrialMetrics`
here, and writes one workbook via :func:`write_group_report`. Every
detected stance phase is kept as explicit touchdown/liftoff frame indices
(see :class:`PawEvents`) rather than folded into an average, so a
validation-video auditor can overlay exactly the frames treated as
ground contact.
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from alligaitor.config import CAMERA_ROLES, GaitConfig
from alligaitor.inference import PoseTrack2D

PathLike = Union[str, Path]

PAW_NODES = ("left-forepaw", "right-forepaw", "left-hind-paw", "right-hind-paw")

CONTRALATERAL = {
    "left-forepaw": "right-forepaw",
    "right-forepaw": "left-forepaw",
    "left-hind-paw": "right-hind-paw",
    "right-hind-paw": "left-hind-paw",
}

# The side camera on the opposite side of the body from a given paw
# (e.g. "right" for left-forepaw), excluded from
# :func:`find_camera_caused_discards`'s attribution since the rat's own
# body occludes that view by construction.
FAR_SIDE_CAMERA = {
    "left-forepaw": "right",
    "right-forepaw": "left",
    "left-hind-paw": "right",
    "right-hind-paw": "left",
}

# Body-center node used to time the crossing and establish the direction
# of travel (see minimal_skeleton.json for the full node set).
REFERENCE_NODE = "mid-back"

# Above this fraction of a paw's usable window coming from the bottom-
# camera fallback (see :mod:`alligaitor.bottom_fallback`), the run is
# still usable but worth a reviewer's second look.
BOTTOM_FALLBACK_WARN_THRESHOLD = 1.0 / 3.0

_INVALID_SHEET_CHARS = set("[]:*?/\\")


@dataclass
class PawEvents:
    """Detected stance (ground-contact) phases for one paw in one trial.

    Each index ``i`` describes one stance phase: the paw is planted from
    ``touchdown_frames[i]``/``touchdown_times[i]`` through
    ``liftoff_frames[i]``/``liftoff_times[i]`` inclusive.
    """

    touchdown_frames: np.ndarray
    liftoff_frames: np.ndarray
    touchdown_times: np.ndarray
    liftoff_times: np.ndarray


@dataclass
class TrialMetrics:
    """Gait metrics for one rat's single crossing of the platform.

    One recording can hold several crossings (see :func:`find_crossings`),
    each its own trial with its own direction of travel.
    ``crossing_index``/``crossing_count`` say which one, and
    ``crossing_window`` is the frame range it was measured over.
    """

    session_name: str
    rat_id: str
    crossing_time_s: float
    average_speed_mm_s: float
    stride_length_mm: Dict[str, float]
    step_length_mm: Dict[str, float]
    ground_contact_time_s: Dict[str, float]
    n_contacts: Dict[str, int]
    n_strides: Dict[str, int]
    n_steps: Dict[str, int]
    paw_events: Dict[str, PawEvents]
    crossing_index: int = 0
    crossing_count: int = 1
    crossing_window: Optional[Tuple[int, int]] = None
    bottom_fallback_fraction: Dict[str, float] = field(default_factory=dict)

    @property
    def crossing_label(self) -> str:
        """How this trial names itself in the workbook -- the bare
        session name when the recording held one crossing, else the
        session name plus which crossing of how many."""
        if self.crossing_count <= 1:
            return self.session_name
        return f"{self.session_name} \u2014 crossing {self.crossing_index + 1} of {self.crossing_count}"


def crossing_block_title(session_name: str, crossing_number: int, crossing_count: int) -> str:
    """The exact title-row text :func:`write_group_report` gives one
    crossing's block in the workbook (mirrors
    :attr:`TrialMetrics.crossing_label`). Used by
    :func:`annotate_manual_flag` to target one crossing's block from
    just the summary JSON's plain numbers."""
    if crossing_count <= 1:
        return f"Session: {session_name}"
    return f"Session: {session_name} \u2014 crossing {crossing_number} of {crossing_count}"


def load_pose_3d(
    csv_path: PathLike,
) -> "tuple[np.ndarray, Dict[str, np.ndarray], Dict[str, np.ndarray], Dict[str, np.ndarray]]":
    """Load a ``pose_3d.csv`` into per-node arrays indexed by frame.

    Returns:
        A ``(times, positions, reprojection_error_px, fallback)`` tuple:
        ``times`` is seconds per frame index; ``positions`` maps node name
        to an ``(n_frames, 3)`` array of that node's ``x``/``y``/``z``,
        ``NaN`` where untriangulated; ``reprojection_error_px`` maps node
        name to an ``(n_frames,)`` array of that frame's mean reprojection
        error (see :class:`alligaitor.triangulation.Pose3D`), ``NaN`` to
        match; ``fallback`` maps node name to an ``(n_frames,)`` boolean
        array, ``True`` wherever that point came from
        :func:`alligaitor.bottom_fallback.fill_gaps` rather than real
        triangulation -- all ``False`` if `csv_path` predates that
        column (see :func:`alligaitor.pipeline.save_pose_3d_csv`).
    """
    df = pd.read_csv(csv_path)
    n_frames = int(df["frame"].max()) + 1
    has_fallback_col = "fallback" in df.columns

    times = np.full(n_frames, np.nan)
    times[df["frame"].to_numpy()] = df["time_s"].to_numpy()

    positions = {}
    reprojection_error_px = {}
    fallback = {}
    for node, sub in df.groupby("node"):
        arr = np.full((n_frames, 3), np.nan)
        arr[sub["frame"].to_numpy()] = sub[["x", "y", "z"]].to_numpy()
        positions[node] = arr

        err = np.full(n_frames, np.nan)
        err[sub["frame"].to_numpy()] = sub["reprojection_error_px"].to_numpy()
        reprojection_error_px[node] = err

        fb = np.zeros(n_frames, dtype=bool)
        if has_fallback_col:
            fb[sub["frame"].to_numpy()] = sub["fallback"].to_numpy(dtype=bool)
        fallback[node] = fb
    return times, positions, reprojection_error_px, fallback


def bridge_short_gaps(xyz: np.ndarray, max_gap: int) -> np.ndarray:
    """Linearly interpolate over short untriangulated runs.

    A run of untriangulated (``NaN``) frames is bridged only when it's at
    most ``max_gap`` frames long *and* bounded by a valid frame on both
    sides -- a gap touching either end of the trial is left as-is, since
    there's nothing to interpolate from/to. ``max_gap <= 0`` is a no-op
    (returns a copy, unchanged).

    Used to keep momentary tracking jitter and brief per-camera dropouts
    (which will always happen, even with better models) from fragmenting
    one real stance phase into pieces too short to individually survive
    :attr:`alligaitor.config.GaitConfig.min_contact_frames` -- see
    :attr:`alligaitor.config.GaitConfig.max_bridge_gap_frames`.
    """
    xyz = xyz.copy()
    if max_gap <= 0:
        return xyz

    valid = ~np.isnan(xyz).any(axis=1)
    n = len(xyz)
    i = 0
    while i < n:
        if valid[i]:
            i += 1
            continue
        j = i
        while j < n and not valid[j]:
            j += 1
        if i > 0 and j < n and (j - i) <= max_gap:
            frac = (np.arange(i, j) - (i - 1)) / (j - (i - 1))
            xyz[i:j] = xyz[i - 1] + (xyz[j] - xyz[i - 1]) * frac[:, None]
        i = j
    return xyz


def pose_sampling_fps(times: np.ndarray) -> float:
    """Frames per second of a ``pose_3d`` time column, recovered from the
    median step between consecutive times so a missing frame can't skew it."""
    dt = np.diff(times)
    dt = dt[np.isfinite(dt) & (dt > 0)]
    if dt.size == 0:
        raise ValueError("pose_3d times contain no usable frame interval; cannot determine fps.")
    return 1.0 / float(np.median(dt))


def windowed_body_speed(times: np.ndarray, xyz: np.ndarray, window_s: float) -> np.ndarray:
    """Per-frame translation speed, in mm/s, measured across a window
    centered on each frame rather than between adjacent frames.

    Frame-to-frame speed of a single triangulated node is dominated by
    reconstruction jitter whenever the animal is at rest, so "has it
    stopped?" can't be asked one frame at a time; net displacement between
    the window's outermost triangulated frames cancels that jitter while
    leaving genuine translation untouched. ``NaN`` on frames whose window
    holds fewer than two triangulated frames, or spans no time at all.
    """
    n = len(xyz)
    half = max(1, int(round(window_s * pose_sampling_fps(times) / 2.0)))
    valid = ~np.isnan(xyz).any(axis=1)
    speed = np.full(n, np.nan)
    for i in range(n):
        lo, hi = max(0, i - half), min(n - 1, i + half)
        idx = np.flatnonzero(valid[lo : hi + 1])
        if idx.size < 2:
            continue
        a, b = lo + int(idx[0]), lo + int(idx[-1])
        span = times[b] - times[a]
        if not span > 0:
            continue
        speed[i] = float(np.linalg.norm(xyz[b] - xyz[a]) / span)
    return speed


def find_crossings(
    times: np.ndarray, ref_xyz: np.ndarray, config: GaitConfig
) -> List[Tuple[int, int]]:
    """Every separate traversal of the platform in one recording, as
    ``(start_frame, end_frame)`` windows in the order they happened.

    An unedited recording typically holds several crossings back to back
    (walk, stop, turn, walk back), each its own trial with its own
    direction of travel, so averaging them together is meaningless.
    Crossings are separated on two signals together: a pause (a stretch
    of at least ``config.min_still_seconds`` below
    ``config.stillness_window_speed_mm_s``) and a reversal (a pause
    followed by travel in the opposite direction) -- a pause alone isn't
    enough, since a rat that stops to groom and continues the same way
    wasn't on a second crossing. A candidate whose net displacement is
    under a quarter of the longest candidate's is dropped as shuffling
    rather than travel. Always returns at least one window, falling back
    to the full triangulated range.
    """
    bridged = bridge_short_gaps(ref_xyz, config.max_bridge_gap_frames)
    valid = ~np.isnan(bridged).any(axis=1)
    valid_idx = np.flatnonzero(valid)
    if valid_idx.size < 2:
        raise ValueError(
            f"Reference node '{REFERENCE_NODE}' has fewer than 2 triangulated frames; "
            "cannot determine an active window."
        )
    first, last = int(valid_idx[0]), int(valid_idx[-1])

    speed = windowed_body_speed(times, bridged, config.stillness_window_seconds)
    with np.errstate(invalid="ignore"):
        still = speed < config.stillness_window_speed_mm_s  # NaN -> moving, never grounds for discarding
    min_still_frames = max(1, int(round(config.min_still_seconds * pose_sampling_fps(times))))

    # Maximal runs of moving frames inside the triangulated range, split by
    # pauses that are actually long enough to be a stop rather than a stride's
    # own slow moment.
    bouts: List[Tuple[int, int]] = []
    i = first
    while i <= last:
        if still[i]:
            i += 1
            continue
        j = i
        while j < last:
            if not still[j + 1]:
                j += 1
                continue
            k = j + 1
            while k <= last and still[k]:
                k += 1
            if k - (j + 1) >= min_still_frames or k > last:
                break  # a real stop (or the end) -- this bout ends at j
            j = k      # too brief to be a stop; keep going through it
        bouts.append((i, j))
        i = j + 1

    if not bouts:
        return [(first, last)]

    def _net(bout: Tuple[int, int]) -> np.ndarray:
        a, b = bout
        seg = np.flatnonzero(valid[a : b + 1])
        if seg.size < 2:
            return np.zeros(3)
        return bridged[a + seg[-1]] - bridged[a + seg[0]]

    # Merge across a pause the rat walked straight out of in the same
    # direction: that was one traversal with a rest in it, not two.
    merged: List[Tuple[int, int]] = [bouts[0]]
    for bout in bouts[1:]:
        prev_net, this_net = _net(merged[-1]), _net(bout)
        if float(prev_net @ this_net) > 0:
            merged[-1] = (merged[-1][0], bout[1])
        else:
            merged.append(bout)

    distances = [float(np.linalg.norm(_net(b))) for b in merged]
    longest = max(distances)
    if longest <= 0:
        return [(first, last)]
    crossings = [b for b, d in zip(merged, distances) if d >= 0.25 * longest]
    return crossings or [(first, last)]


def active_window(times: np.ndarray, ref_xyz: np.ndarray, config: GaitConfig) -> Tuple[int, int]:
    """First/last frame index of sustained whole-body motion, for the
    *first* crossing in the recording (delegates to :func:`find_crossings`).

    Trims any leading/trailing stretch of at least
    ``config.min_still_seconds`` whose windowed body speed (see
    :func:`windowed_body_speed`) stays below
    ``config.stillness_window_speed_mm_s``, so idle time before/after
    motion doesn't look like real strides. A frame with no windowed speed
    counts as moving, not still.
    """
    return find_crossings(times, ref_xyz, config)[0]


def _crossing_time_and_speed(
    times: np.ndarray,
    ref_xyz: np.ndarray,
    config: GaitConfig,
    window: Optional[Tuple[int, int]] = None,
) -> "tuple[float, float, np.ndarray, Tuple[int, int]]":
    """Return (crossing time, average speed, unit forward direction, active window).

    Restricted to the active window (see :func:`active_window`) so idle
    time before/after motion doesn't inflate crossing time or drag down
    average speed. ``window`` names one crossing's ``(start, end)``
    explicitly for a recording holding more than one; omitted, the first
    crossing is used.
    """
    start, end = window if window is not None else active_window(times, ref_xyz, config)

    window = ref_xyz[start : end + 1]
    valid = ~np.isnan(window).any(axis=1)
    valid_idx = np.flatnonzero(valid)
    if valid_idx.size < 2:
        raise ValueError(
            f"Reference node '{REFERENCE_NODE}' has fewer than 2 triangulated frames "
            "in its active window; cannot determine crossing time or direction."
        )

    first_idx, last_idx = start + valid_idx[0], start + valid_idx[-1]
    net_displacement = ref_xyz[last_idx] - ref_xyz[first_idx]
    net_distance = np.linalg.norm(net_displacement)
    if net_distance == 0:
        raise ValueError(
            f"Reference node '{REFERENCE_NODE}' shows no net displacement; "
            "cannot determine crossing direction."
        )
    forward = net_displacement / net_distance

    crossing_time_s = times[last_idx] - times[first_idx]

    pts = ref_xyz[start + valid_idx]
    path_length = np.linalg.norm(np.diff(pts, axis=0), axis=1).sum()
    average_speed_mm_s = path_length / crossing_time_s

    return float(crossing_time_s), float(average_speed_mm_s), forward, (start, end)


def _raw_stance_candidates(
    times: np.ndarray, xyz: np.ndarray, config: GaitConfig
) -> "tuple[np.ndarray, np.ndarray, list[tuple[int, int]]]":
    """Shared groundwork for stance detection.

    Returns ``(valid, speed, runs)``: whether the paw is triangulated on
    each frame, frame-to-frame speed, and every maximal run of raw
    "planted" frames (speed below ``config.speed_threshold_mm_s``), before
    the ``min_contact_frames`` length filter. Used by both
    :func:`_detect_paw_events` and :func:`find_camera_caused_discards`.

    Speed at each frame is the *minimum* of its backward- and
    forward-difference estimates (whichever neighbor is triangulated), not
    backward-only, since a frame right after an untriangulated gap would
    otherwise have its backward difference span the whole gap and inflate
    apparent speed at a real stance's onset.
    """
    n = len(times)
    valid = ~np.isnan(xyz).any(axis=1)

    disp = np.linalg.norm(xyz[1:] - xyz[:-1], axis=1)
    dt = times[1:] - times[:-1]
    with np.errstate(invalid="ignore", divide="ignore"):
        step_speed = disp / dt  # step_speed[k] = speed of the transition frame k -> frame k+1

    backward = np.full(n, np.nan)
    backward[1:] = step_speed
    forward = np.full(n, np.nan)
    forward[:-1] = step_speed
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)  # all-NaN slice at isolated valid frames
        speed = np.nanmin(np.stack([backward, forward]), axis=0)

    planted = valid & (speed < config.speed_threshold_mm_s)

    runs = []
    start = None
    for i, p in enumerate(planted):
        if p and start is None:
            start = i
        elif not p and start is not None:
            runs.append((start, i - 1))
            start = None
    if start is not None:
        runs.append((start, n - 1))

    return valid, speed, runs


def _drop_window_clipped_stances(events: PawEvents, start: int, end: int) -> PawEvents:
    """Drop stance phases whose touchdown or liftoff is the active
    window's boundary rather than a detected event.

    Paw positions are masked to the active window (see
    :func:`active_window`) before stance detection, so a paw already
    planted when the window opens, or still planted when it closes,
    yields a stance phase with a synthetic endpoint (the mask edge, not a
    real touchdown/liftoff) that would otherwise inflate ground contact
    time.
    """
    keep = (events.touchdown_frames > start) & (events.liftoff_frames < end)
    return PawEvents(
        touchdown_frames=events.touchdown_frames[keep],
        liftoff_frames=events.liftoff_frames[keep],
        touchdown_times=events.touchdown_times[keep],
        liftoff_times=events.liftoff_times[keep],
    )


def _detect_paw_events(times: np.ndarray, xyz: np.ndarray, config: GaitConfig) -> PawEvents:
    """Segment one paw's trajectory into stance phases.

    A frame counts as planted when the paw is triangulated and its
    frame-to-frame speed (see :func:`_raw_stance_candidates`) is below
    ``config.speed_threshold_mm_s``. Runs of fewer than
    ``config.min_contact_frames`` planted frames are dropped as tracking
    jitter rather than counted as real stance phases -- see
    :func:`find_camera_caused_discards` for identifying which of those
    discards trace back to a specific camera's dropped detection.
    """
    empty = PawEvents(np.array([], dtype=int), np.array([], dtype=int), np.array([]), np.array([]))

    valid, _, runs = _raw_stance_candidates(times, xyz, config)
    if not valid.any():
        return empty

    runs = [(s, e) for s, e in runs if (e - s + 1) >= config.min_contact_frames]
    if not runs:
        return empty

    touchdown_frames = np.array([s for s, _ in runs], dtype=int)
    liftoff_frames = np.array([e for _, e in runs], dtype=int)
    return PawEvents(
        touchdown_frames=touchdown_frames,
        liftoff_frames=liftoff_frames,
        touchdown_times=times[touchdown_frames],
        liftoff_times=times[liftoff_frames],
    )


@dataclass
class DiscardedStance:
    """An untriangulated gap that could plausibly be hiding a real stance.

    Attributes:
        start_frame: First frame of the gap.
        end_frame: Last frame of the gap.
        dropped_by: Camera role(s) missing a valid detection somewhere
            in the gap.
    """

    start_frame: int
    end_frame: int
    dropped_by: List[str]


def find_camera_caused_discards(
    bridged_xyz: np.ndarray,
    cam_valid: Dict[str, np.ndarray],
    exclude_camera: Optional[str] = None,
) -> List[DiscardedStance]:
    """Find every untriangulated gap bounded by a triangulated frame on
    both sides, and which camera(s) -- other than ``exclude_camera`` --
    were missing somewhere in it.

    ``bridged_xyz`` should already have short gaps interpolated (see
    :func:`bridge_short_gaps`), so what's left here is exactly the
    gaps long/unresolved enough to matter. A gap touching either end of
    the trial isn't reported -- there's no before/after stance it could
    plausibly be separating.

    Args:
        bridged_xyz: One paw's ``(n_frames, 3)`` bridged positions.
        cam_valid: Per camera role, an ``(n_frames,)`` boolean array --
            whether that camera had a valid (aligned) 2D detection for
            this paw on each frame. See
            :func:`alligaitor.triangulation.align_tracks_by_time`.
        exclude_camera: A camera role never counted as having dropped the
            paw -- e.g. :data:`FAR_SIDE_CAMERA`\\ [paw], since that
            camera not seeing this paw is expected, not a failure. A
            genuine gap always has at least one non-excluded camera
            missing somewhere in it: triangulation needs >=2 valid
            cameras, so if both non-excluded cameras were valid
            throughout, there'd be no gap to find in the first place.
    """
    valid = ~np.isnan(bridged_xyz).any(axis=1)
    n = len(valid)

    def _cameras_missing_at(f: int) -> "set[str]":
        return {role for role, cv in cam_valid.items() if role != exclude_camera and not cv[f]}

    results = []
    i = 0
    while i < n:
        if valid[i]:
            i += 1
            continue
        j = i
        while j < n and not valid[j]:
            j += 1
        if i > 0 and j < n:  # interior gap only -- skip leading/trailing edges
            dropped_by = set()
            for f in range(i, j):
                dropped_by |= _cameras_missing_at(f)
            if dropped_by:
                results.append(DiscardedStance(i, j - 1, sorted(dropped_by)))
        i = j

    return results


def cam_valid_by_paw_from_aligned(aligned: Dict[str, PoseTrack2D]) -> Dict[str, Dict[str, np.ndarray]]:
    """Per paw, per camera role, whether that camera had a valid 2D
    detection on each shared-timeline frame -- the ``cam_valid`` input
    :func:`find_camera_caused_discards` needs, derived from tracks
    already resampled onto the shared timeline (see
    :func:`alligaitor.triangulation.align_tracks_by_time`).
    """
    cam_valid_by_paw = {}
    for paw in PAW_NODES:
        node_idx = {role: aligned[role].node_names.index(paw) for role in CAMERA_ROLES}
        cam_valid_by_paw[paw] = {
            role: ~np.isnan(aligned[role].points[:, node_idx[role], :]).any(axis=1) for role in CAMERA_ROLES
        }
    return cam_valid_by_paw


def compute_discards_by_paw(
    positions: Dict[str, np.ndarray],
    cam_valid_by_paw: Dict[str, Dict[str, np.ndarray]],
    config: GaitConfig,
) -> Dict[str, List[DiscardedStance]]:
    """Per paw, its camera-caused discarded gaps (see
    :func:`find_camera_caused_discards`), bridging short gaps first to
    match what :func:`compute_trial_metrics` itself sees."""
    return {
        paw: find_camera_caused_discards(
            bridge_short_gaps(positions[paw], config.max_bridge_gap_frames),
            cam_valid_by_paw[paw],
            exclude_camera=FAR_SIDE_CAMERA[paw],
        )
        for paw in PAW_NODES
    }


def find_stride_length_outliers(
    events: PawEvents,
    xyz: np.ndarray,
    forward: np.ndarray,
    ratio: float,
) -> List[Tuple[int, int, float, float]]:
    """Adjacent stride pairs that aren't trustworthy as one real stride:
    non-positive (the paw's own net forward progress, liftoff to next
    touchdown, is zero or backward), or long enough to exceed ``ratio``
    times this paw's own median *positive* stride length in the trial.

    A too-long stride usually means a real stance sat in between that the
    speed classifier missed -- complementary to
    :func:`find_camera_caused_discards`'s gap-based check. A non-positive
    stride means the paw was genuinely tracked moving with no net forward
    progress (e.g. a mid-pause the speed classifier still called a
    "swing"); sign alone flags it, since directed locomotion should never
    produce a non-positive stride. The baseline median is computed from
    positive strides only, so a paused/reversed stride can't mask a
    genuine too-long outlier.

    Returns a list of ``(liftoff_frame, touchdown_frame, stride_length_mm,
    median_stride_length_mm)``, one entry per outlier stride.
    ``median_stride_length_mm`` is ``NaN`` when fewer than two positive
    strides exist to compute one from.
    """
    strides = _stride_lengths(events, xyz, forward)
    if strides.size == 0:
        return []
    non_positive = strides <= 0
    positive = strides[~non_positive]
    if positive.size >= 2:
        median = float(np.median(positive))
        too_long = (~non_positive) & (strides > median * ratio)
    else:
        median = float("nan")
        too_long = np.zeros(strides.size, dtype=bool)
    is_outlier = non_positive | too_long
    return [
        (int(events.liftoff_frames[i]), int(events.touchdown_frames[i + 1]), float(s), median)
        for i, s in enumerate(strides)
        if is_outlier[i]
    ]


def _qualifying_runs(
    events: PawEvents,
    bridged_xyz: np.ndarray,
    forward: np.ndarray,
    min_run: int,
    stride_length_outlier_ratio: float,
) -> List[Tuple[int, int]]:
    """Maximal runs of consecutive accepted stance events, as ``(start, end)``
    index pairs into ``events``' arrays, at least ``min_run`` events long.

    A pair of adjacent events stays in the same run only if *both* hold:
    no remaining untriangulated frame (after bridging, see
    :func:`bridge_short_gaps`) in the swing between them, and the stride
    between them isn't a :func:`find_stride_length_outliers`-flagged
    outlier. The first catches a triangulation gap that could be hiding a
    real stance; the second catches a real stance missed despite clean
    triangulation. Neither alone is sufficient -- see
    :func:`restrict_to_consecutive_runs`.
    """
    n = events.touchdown_frames.size
    if n == 0:
        return []

    valid = ~np.isnan(bridged_xyz).any(axis=1)
    gap_clean = np.ones(max(n - 1, 0), dtype=bool)
    for i in range(n - 1):
        gap_start, gap_end = events.liftoff_frames[i], events.touchdown_frames[i + 1]
        gap_clean[i] = bool(valid[gap_start : gap_end + 1].all())

    strides = _stride_lengths(events, bridged_xyz, forward)
    stride_clean = np.ones(max(n - 1, 0), dtype=bool)
    if strides.size:
        non_positive = strides <= 0
        positive = strides[~non_positive]
        too_long = np.zeros(strides.size, dtype=bool)
        if positive.size >= 2:
            median = float(np.median(positive))
            too_long = (~non_positive) & (strides > median * stride_length_outlier_ratio)
        stride_clean = ~(non_positive | too_long)

    clean_pair = gap_clean & stride_clean

    runs = []
    start = 0
    for i in range(n - 1):
        if not clean_pair[i]:
            runs.append((start, i))
            start = i + 1
    runs.append((start, n - 1))

    return [(s, e) for s, e in runs if (e - s + 1) >= min_run]


def restrict_to_consecutive_runs(
    trial: TrialMetrics,
    times: np.ndarray,
    positions: Dict[str, np.ndarray],
    config: GaitConfig,
    window: Optional[Tuple[int, int]] = None,
) -> TrialMetrics:
    """Recompute stride/step/ground-contact-time averages from only the
    strides that are part of a qualifying run (see
    :attr:`alligaitor.config.GaitConfig.min_consecutive_strides` and
    :func:`_qualifying_runs`); a paw with no qualifying run reports
    ``NaN``. ``trial.paw_events`` is left untouched, so the validation
    video and raw event log still show everything treated as ground
    contact -- only the summary numbers are restricted.

    Args:
        trial: Already-computed metrics (see :func:`compute_trial_metrics`).
        times: This trial's per-frame timestamps (see :func:`load_pose_3d`).
        positions: This trial's per-node positions (see :func:`load_pose_3d`).
        config: The same :class:`GaitConfig` ``trial`` was computed with.
    """
    _, _, forward, (start, end) = _crossing_time_and_speed(
        times, positions[REFERENCE_NODE], config, window or trial.crossing_window
    )

    stride_length_mm, step_length_mm, ground_contact_time_s = {}, {}, {}
    n_contacts, n_strides, n_steps = {}, {}, {}

    for paw in PAW_NODES:
        events = trial.paw_events[paw]
        contra_events = trial.paw_events[CONTRALATERAL[paw]]
        xyz = positions[paw].copy()
        xyz[:start] = np.nan
        xyz[end + 1 :] = np.nan
        bridged = bridge_short_gaps(xyz, config.max_bridge_gap_frames)
        contra_xyz = positions[CONTRALATERAL[paw]].copy()
        contra_xyz[:start] = np.nan
        contra_xyz[end + 1 :] = np.nan
        contra_bridged = bridge_short_gaps(contra_xyz, config.max_bridge_gap_frames)
        runs = _qualifying_runs(
            events, bridged, forward, config.min_consecutive_strides, config.stride_length_outlier_ratio
        )

        if not runs:
            stride_length_mm[paw] = float("nan")
            step_length_mm[paw] = float("nan")
            ground_contact_time_s[paw] = float("nan")
            n_contacts[paw] = 0
            n_strides[paw] = 0
            n_steps[paw] = 0
            continue

        strides_parts, steps_parts, durations_parts = [], [], []
        total_contacts = 0
        for s, e in runs:
            run_events = PawEvents(
                touchdown_frames=events.touchdown_frames[s : e + 1],
                liftoff_frames=events.liftoff_frames[s : e + 1],
                touchdown_times=events.touchdown_times[s : e + 1],
                liftoff_times=events.liftoff_times[s : e + 1],
            )
            # Use `bridged`, not raw `positions`: a touchdown landing on a
            # bridged frame reads NaN out of the raw array, which would
            # turn this paw's whole mean into NaN.
            strides_parts.append(_stride_lengths(run_events, bridged, forward))
            steps_parts.append(
                _step_lengths(
                    run_events, contra_events, bridged, contra_bridged, forward,
                    config.stride_length_outlier_ratio,
                )
            )
            durations_parts.append(run_events.liftoff_times - run_events.touchdown_times)
            total_contacts += run_events.touchdown_frames.size

        strides = np.concatenate(strides_parts)
        steps = np.concatenate(steps_parts)
        durations = np.concatenate(durations_parts)

        stride_length_mm[paw] = float(np.mean(strides)) if strides.size else float("nan")
        # Step length carries its own evidence bar (config.min_valid_steps): it
        # is the only metric depending on a second paw, so it can fall short on
        # a crossing where this paw's own stride/contact numbers are sound.
        step_length_mm[paw] = (
            float(np.mean(steps)) if steps.size >= config.min_valid_steps else float("nan")
        )
        ground_contact_time_s[paw] = float(np.mean(durations)) if durations.size else float("nan")
        n_contacts[paw] = total_contacts
        n_strides[paw] = int(strides.size)
        n_steps[paw] = int(steps.size)

    return replace(
        trial,
        stride_length_mm=stride_length_mm,
        step_length_mm=step_length_mm,
        ground_contact_time_s=ground_contact_time_s,
        n_contacts=n_contacts,
        n_strides=n_strides,
        n_steps=n_steps,
    )


def _stride_lengths(events: PawEvents, xyz: np.ndarray, forward: np.ndarray) -> np.ndarray:
    """Forward distance from each stance phase's liftoff to the next touchdown."""
    if events.touchdown_frames.size < 2:
        return np.array([])
    liftoff_pos = xyz[events.liftoff_frames[:-1]]
    touchdown_pos = xyz[events.touchdown_frames[1:]]
    return (touchdown_pos - liftoff_pos) @ forward


def _step_lengths(
    events: PawEvents,
    contra_events: PawEvents,
    xyz: np.ndarray,
    contra_xyz: np.ndarray,
    forward: np.ndarray,
    outlier_ratio: float,
) -> np.ndarray:
    """Forward distance from each touchdown to the contralateral paw's
    most recent touchdown *strictly before* it -- keeping only the
    pairings that contralateral paw's own record supports.

    "Strictly before," not at-or-before, since two paws' touchdowns can
    legitimately land in the same discretized frame (a real
    double-support moment), and step length isn't meaningful between two
    simultaneous footfalls. A pairing is rejected when the contralateral
    touchdown interval bracketing it runs longer than ``outlier_ratio``
    times that paw's own median touchdown-to-touchdown interval -- the
    same baseline-relative test :func:`find_stride_length_outliers`
    applies to strides, so a consistently-impaired paw with a long but
    stable interval isn't flagged. Untriangulated dropouts aren't treated
    as invalidating on their own, since a dropout that does hide a
    touchdown already stretches the bracketing interval past the limit.

    Returns an empty array when the contralateral paw has fewer than two
    touchdowns.
    """
    if events.touchdown_frames.size == 0 or contra_events.touchdown_frames.size < 2:
        return np.array([])

    contra_idx = np.searchsorted(contra_events.touchdown_frames, events.touchdown_frames, side="left") - 1
    has_prior = contra_idx >= 0
    if not has_prior.any():
        return np.array([])

    contra_intervals = np.diff(contra_events.touchdown_frames)
    limit = outlier_ratio * float(np.median(contra_intervals))

    this_frames = events.touchdown_frames[has_prior]
    idx = contra_idx[has_prior]
    contra_frames = contra_events.touchdown_frames[idx]

    # The contralateral interval bracketing each pairing: that touchdown to
    # its own next one, or -- for a pairing after the contralateral paw's
    # last touchdown -- however far it is to this paw's touchdown.
    nxt = idx + 1
    has_next = nxt < contra_events.touchdown_frames.size
    span = np.where(
        has_next,
        contra_events.touchdown_frames[np.minimum(nxt, contra_events.touchdown_frames.size - 1)] - contra_frames,
        this_frames - contra_frames,
    )
    keep = span <= limit
    if not keep.any():
        return np.array([])

    this_pos = xyz[this_frames[keep]]
    contra_pos = contra_xyz[contra_frames[keep]]
    return (this_pos - contra_pos) @ forward


def compute_trial_metrics(
    csv_path: PathLike,
    session_name: str,
    rat_id: str,
    config: Optional[GaitConfig] = None,
    window: Optional[Tuple[int, int]] = None,
    crossing_index: int = 0,
    crossing_count: int = 1,
) -> TrialMetrics:
    """Compute one trial's gait metrics from its triangulated ``pose_3d.csv``.

    Args:
        csv_path: This trial's ``pose_3d.csv``, as written by
            :func:`alligaitor.pipeline.run_session`.
        session_name: Trial identifier, carried through to the report.
        rat_id: Which rat this trial belongs to (see
            :attr:`alligaitor.config.SessionConfig.rat_id`); sessions
            sharing a ``rat_id`` land on the same spreadsheet tab.
        config: Stance/swing detection thresholds; defaults to
            :class:`GaitConfig`'s own defaults.
    """
    config = config or GaitConfig()
    times, positions, _, _ = load_pose_3d(csv_path)

    missing = [node for node in (REFERENCE_NODE, *PAW_NODES) if node not in positions]
    if missing:
        raise ValueError(f"pose_3d CSV '{csv_path}' is missing required node(s): {missing}")

    crossing_time_s, average_speed_mm_s, forward, (start, end) = _crossing_time_and_speed(
        times, positions[REFERENCE_NODE], config, window
    )

    # Bridged (short-gap-interpolated) paw positions drive stance
    # detection and every downstream paw measurement, masked to the
    # active window first so a paw jittering after the rat stops can't
    # be detected as a stance phase.
    bridged = {}
    for paw in PAW_NODES:
        xyz = positions[paw].copy()
        xyz[:start] = np.nan
        xyz[end + 1 :] = np.nan
        bridged[paw] = bridge_short_gaps(xyz, config.max_bridge_gap_frames)

    # A stance touching either edge of the active window was cut by the
    # masking above rather than by a real liftoff/touchdown, so it isn't
    # a completed ground contact -- see _drop_window_clipped_stances.
    events = {
        paw: _drop_window_clipped_stances(
            _detect_paw_events(times, bridged[paw], config), start, end
        )
        for paw in PAW_NODES
    }

    stride_length_mm, step_length_mm, ground_contact_time_s = {}, {}, {}
    n_contacts, n_strides, n_steps = {}, {}, {}

    for paw in PAW_NODES:
        ev = events[paw]
        contra_ev = events[CONTRALATERAL[paw]]

        strides = _stride_lengths(ev, bridged[paw], forward)
        steps = _step_lengths(
            ev, contra_ev, bridged[paw], bridged[CONTRALATERAL[paw]], forward,
            config.stride_length_outlier_ratio,
        )
        contact_durations = ev.liftoff_times - ev.touchdown_times

        stride_length_mm[paw] = float(np.mean(strides)) if strides.size else float("nan")
        step_length_mm[paw] = (
            float(np.mean(steps)) if steps.size >= config.min_valid_steps else float("nan")
        )
        ground_contact_time_s[paw] = float(np.mean(contact_durations)) if contact_durations.size else float("nan")
        n_contacts[paw] = int(ev.touchdown_frames.size)
        n_strides[paw] = int(strides.size)
        n_steps[paw] = int(steps.size)

    return TrialMetrics(
        session_name=session_name,
        rat_id=rat_id,
        crossing_index=crossing_index,
        crossing_count=crossing_count,
        crossing_window=(start, end),
        crossing_time_s=crossing_time_s,
        average_speed_mm_s=average_speed_mm_s,
        stride_length_mm=stride_length_mm,
        step_length_mm=step_length_mm,
        ground_contact_time_s=ground_contact_time_s,
        n_contacts=n_contacts,
        n_strides=n_strides,
        n_steps=n_steps,
        paw_events=events,
    )


def compute_crossing_metrics(
    csv_path: PathLike,
    session_name: str,
    rat_id: str,
    config: Optional[GaitConfig] = None,
) -> List[TrialMetrics]:
    """One :class:`TrialMetrics` per crossing in a recording (see
    :func:`find_crossings`), each measured independently with its own
    active window and direction of travel. All share ``rat_id``, so they
    land on one worksheet tab with a combined average (see
    :func:`write_group_report`).
    """
    config = config or GaitConfig()
    times, positions, _, _ = load_pose_3d(csv_path)
    if REFERENCE_NODE not in positions:
        raise ValueError(f"pose_3d CSV '{csv_path}' is missing required node: {REFERENCE_NODE}")

    crossings = find_crossings(times, positions[REFERENCE_NODE], config)
    trials = []
    for i, window in enumerate(crossings):
        try:
            trials.append(
                compute_trial_metrics(
                    csv_path, session_name, rat_id, config,
                    window=window, crossing_index=i, crossing_count=len(crossings),
                )
            )
        except ValueError as exc:
            # One unusable crossing shouldn't cost the others.
            warnings.warn(
                f"{session_name}: skipping crossing {i + 1} of {len(crossings)} "
                f"(frames {window[0]}-{window[1]}): {exc}",
                stacklevel=2,
            )
    return trials


@dataclass
class PawWindow:
    """The single time window a validation-video viewer should highlight
    for one paw: its longest qualifying run (see :func:`_qualifying_runs`)
    if it has one, else its longest raw detected run regardless of
    length. ``usable`` says which case applies.
    """

    start_frame: int
    end_frame: int
    start_s: float
    end_s: float
    duration_s: float
    usable: bool
    bottom_fallback_fraction: float = 0.0


def _longest_raw_run(events: PawEvents) -> Optional[Tuple[int, int]]:
    """Index range (into `events`' arrays) of the longest single detected
    stance phase, or ``None`` if nothing was ever detected. Used as the
    fallback window for a paw with zero qualifying runs."""
    n = events.touchdown_frames.size
    if n == 0:
        return None
    durations = events.liftoff_times - events.touchdown_times
    i = int(np.argmax(durations))
    return i, i


def paw_usability_windows(
    trial: TrialMetrics,
    times: np.ndarray,
    positions: Dict[str, np.ndarray],
    config: GaitConfig,
    bottom_fallback_mask: Optional[Dict[str, np.ndarray]] = None,
) -> Dict[str, Optional[PawWindow]]:
    """Per paw, the single window a validation-video viewer should
    highlight: the longest qualifying run (see :func:`_qualifying_runs`)
    if the paw has one, else its longest raw detected stance phase.
    ``None`` only when the paw was never detected planted in this trial.

    Args:
        bottom_fallback_mask: Per paw, an ``(n_frames,)`` boolean array
            (see :func:`load_pose_3d`'s ``fallback`` return value), used
            to compute the window's ``bottom_fallback_fraction``. ``None``
            (the default) leaves every window's fraction at ``0.0``.
    """
    _, _, forward, (start, end) = _crossing_time_and_speed(
        times, positions[REFERENCE_NODE], config, trial.crossing_window
    )

    windows: Dict[str, Optional[PawWindow]] = {}
    for paw in PAW_NODES:
        events = trial.paw_events[paw]
        xyz = positions[paw].copy()
        xyz[:start] = np.nan
        xyz[end + 1 :] = np.nan
        bridged = bridge_short_gaps(xyz, config.max_bridge_gap_frames)
        runs = _qualifying_runs(
            events, bridged, forward, config.min_consecutive_strides, config.stride_length_outlier_ratio
        )

        if runs:
            s, e = max(runs, key=lambda r: events.liftoff_times[r[1]] - events.touchdown_times[r[0]])
            usable = True
        else:
            fallback = _longest_raw_run(events)
            if fallback is None:
                windows[paw] = None
                continue
            s, e = fallback
            usable = False

        start_frame, end_frame = int(events.touchdown_frames[s]), int(events.liftoff_frames[e])
        bottom_fallback_fraction = 0.0
        if bottom_fallback_mask is not None and paw in bottom_fallback_mask:
            segment = bottom_fallback_mask[paw][start_frame : end_frame + 1]
            if segment.size:
                bottom_fallback_fraction = float(segment.mean())

        windows[paw] = PawWindow(
            start_frame=start_frame,
            end_frame=end_frame,
            start_s=float(events.touchdown_times[s]),
            end_s=float(events.liftoff_times[e]),
            duration_s=float(events.liftoff_times[e] - events.touchdown_times[s]),
            usable=usable,
            bottom_fallback_fraction=bottom_fallback_fraction,
        )
    return windows


def attach_bottom_fallback_fraction(
    trial: TrialMetrics,
    times: np.ndarray,
    positions: Dict[str, np.ndarray],
    bottom_fallback_mask: Dict[str, np.ndarray],
    config: GaitConfig,
) -> TrialMetrics:
    """Return `trial` with :attr:`TrialMetrics.bottom_fallback_fraction`
    populated: per paw, how much of its usable window (see
    :func:`paw_usability_windows`) came from
    :func:`alligaitor.bottom_fallback.fill_gaps` rather than real
    triangulation. A paw with no window at all (never detected planted)
    gets ``0.0`` -- there's no run to have come from the fallback.

    Used by :func:`alligaitor.pipeline.run_group` so
    :func:`write_group_report` can warn on a run that's mostly
    fallback-derived, without recomputing windows itself.
    """
    windows = paw_usability_windows(trial, times, positions, config, bottom_fallback_mask=bottom_fallback_mask)
    fractions = {paw: (w.bottom_fallback_fraction if w is not None else 0.0) for paw, w in windows.items()}
    return replace(trial, bottom_fallback_fraction=fractions)


def planted_mask(trial: TrialMetrics, n_frames: int) -> Dict[str, np.ndarray]:
    """Per-paw boolean array, ``True`` on frames within a detected stance phase.

    Expands :attr:`TrialMetrics.paw_events`' touchdown/liftoff frame pairs
    into a full per-frame mask -- e.g. for a validation video to color a
    paw node by its ground-contact state frame by frame.
    """
    masks = {}
    for paw in PAW_NODES:
        mask = np.zeros(n_frames, dtype=bool)
        ev = trial.paw_events[paw]
        for touchdown, liftoff in zip(ev.touchdown_frames, ev.liftoff_frames):
            mask[touchdown : liftoff + 1] = True
        masks[paw] = mask
    return masks


def save_paw_events_csv(
    trial: Union[TrialMetrics, List[TrialMetrics]], csv_path: PathLike
) -> None:
    """Write every detected stance phase's frame/time window for one
    recording, across all of its crossings if it holds several. Accepts
    either one trial or the per-crossing list from
    :func:`compute_crossing_metrics`; the ``crossing`` column (1-based)
    says which traversal each stance belongs to.
    """
    trials = [trial] if isinstance(trial, TrialMetrics) else list(trial)
    rows = [
        {
            "crossing": t.crossing_index + 1,
            "paw": paw,
            "touchdown_frame": int(td_f),
            "liftoff_frame": int(lo_f),
            "touchdown_time_s": float(td_t),
            "liftoff_time_s": float(lo_t),
        }
        for t in trials
        for paw, ev in t.paw_events.items()
        for td_f, lo_f, td_t, lo_t in zip(
            ev.touchdown_frames, ev.liftoff_frames, ev.touchdown_times, ev.liftoff_times
        )
    ]
    df = pd.DataFrame(
        rows,
        columns=["crossing", "paw", "touchdown_frame", "liftoff_frame",
                 "touchdown_time_s", "liftoff_time_s"],
    )
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False)


_PAW_LABELS: Dict[str, str] = {
    "left-forepaw": "Left forepaw",
    "right-forepaw": "Right forepaw",
    "left-hind-paw": "Left hind paw",
    "right-hind-paw": "Right hind paw",
}

# (attribute name on TrialMetrics, column header, per-session number format).
# Counts are true per-session integers, so "0" is right; the averages
# table overrides every column to "0.00" instead, since an average of
# counts across crossings is a fraction.
_STAT_COLUMNS: Tuple[Tuple[str, str, str], ...] = (
    ("stride_length_mm", "Stride Length (mm)", "0.00"),
    ("step_length_mm", "Step Length (mm)", "0.00"),
    ("ground_contact_time_s", "Ground Contact Time (s)", "0.00"),
    ("n_contacts", "Contacts (n)", "0"),
    ("n_strides", "Strides (n)", "0"),
    ("n_steps", "Steps (n)", "0"),
)
# The stats that can legitimately be NaN for an otherwise-fine paw; the
# rest are counts (see _paw_has_no_usable_run).
_NAN_BEARING_STATS = frozenset({"stride_length_mm", "step_length_mm", "ground_contact_time_s"})
_NOTES_LABEL = "Notes"
_N_STAT_COLUMNS = 1 + len(_STAT_COLUMNS)  # "Paw" + the stat columns
_NOTES_COLUMN = _N_STAT_COLUMNS + 1
_N_COLUMNS = _NOTES_COLUMN  # "Paw" + the stat columns + Notes

_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
_TITLE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_LABEL_FONT = Font(bold=True)
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_BAD_FONT = Font(color="9C0006")
# A run that's still usable but leans heavily on the experimental
# bottom-camera fallback (see BOTTOM_FALLBACK_WARN_THRESHOLD) -- Excel's
# standard "Neutral" yellow, deliberately distinct from _BAD_FILL's red so
# a reviewer doesn't mistake a usable-but-caution run for an unusable one.
_WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_WARN_FONT = Font(color="9C6500")
_CLEAR_FILL = PatternFill(fill_type=None)
_CLEAR_FONT = Font()


def _paw_has_no_usable_run(trial: TrialMetrics, paw: str) -> bool:
    """True if `paw` never formed a run clean/long enough to trust for
    stride length, step length, or ground-contact time on this crossing
    (all three NaN). ``n_contacts`` et al. can still be a real, nonzero
    count even when this is True."""
    return (
        np.isnan(trial.stride_length_mm[paw])
        and np.isnan(trial.step_length_mm[paw])
        and np.isnan(trial.ground_contact_time_s[paw])
    )


def _stat_is_untrustworthy(trial: TrialMetrics, paw: str, stat: str) -> bool:
    """True if this one ``(paw, stat)`` cell shouldn't be trusted, as
    opposed to the whole paw (:func:`_paw_has_no_usable_run`).

    Step length can fail on its own since it's the only metric measured
    against a second paw (see
    :attr:`alligaitor.config.GaitConfig.min_valid_steps`), so a crossing
    can be otherwise sound but still have too few trustworthy step
    pairings. Event counts are always a real count, never untrustworthy
    on their own.
    """
    if stat not in _NAN_BEARING_STATS:
        return False
    return bool(np.isnan(getattr(trial, stat)[paw]))


def _nanmean(values) -> float:
    with warnings.catch_warnings():
        # An all-NaN slice (a paw with zero qualifying detections across
        # every crossing) warns by default -- NaN is exactly the right
        # answer there, not a bug worth surfacing.
        warnings.simplefilter("ignore", category=RuntimeWarning)
        return float(np.nanmean(values))


def _write_cell(ws, row: int, col: int, value, number_format: str, bad: bool):
    display_value = None if value is None or (isinstance(value, float) and np.isnan(value)) else value
    cell = ws.cell(row=row, column=col, value=display_value)
    cell.number_format = number_format
    if bad:
        cell.fill = _BAD_FILL
        cell.font = _BAD_FONT
    return cell


def _write_paw_block(
    ws,
    start_row: int,
    title: str,
    crossing_time_s: float,
    average_speed_cm_s: float,
    get_value: Callable[[str, str], float],
    is_bad: Callable[[str, str], bool],
    stat_columns: Tuple[Tuple[str, str, str], ...] = _STAT_COLUMNS,
    note_getter: Callable[[str], str] = lambda paw: "",
    is_warn: Callable[[str], bool] = lambda paw: False,
) -> int:
    """Writes one titled block -- crossing time/speed, then a paw x stat
    table -- starting at `start_row`, and returns the row the next block
    should start at. `get_value(stat, paw)` and `is_bad(paw, stat)`
    abstract over "one trial's own numbers" vs. "this rat's per-paw
    averages" (see write_group_report). `is_bad` is per ``(paw, stat)``
    so a single untrustworthy column is highlighted on its own instead of
    the whole row. `is_warn(paw)` marks a paw yellow instead of red (a
    usable run worth a second look), checked only when not already
    `is_bad`.
    """
    row = start_row
    title_cell = ws.cell(row=row, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.fill = _TITLE_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLUMNS)
    for col in range(1, _N_COLUMNS + 1):
        ws.cell(row=row, column=col).fill = _TITLE_FILL
    row += 1

    ws.cell(row=row, column=1, value="Crossing Time (s)").font = _LABEL_FONT
    _write_cell(ws, row, 2, crossing_time_s, "0.00", bad=False)
    ws.cell(row=row, column=3, value="Average Speed (cm/s)").font = _LABEL_FONT
    _write_cell(ws, row, 4, average_speed_cm_s, "0.00", bad=False)
    row += 2  # blank spacer before the paw table

    ws.cell(row=row, column=1, value="Paw").font = _HEADER_FONT
    ws.cell(row=row, column=1).fill = _HEADER_FILL
    for col, (_, label, _fmt) in enumerate(stat_columns, start=2):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    notes_header = ws.cell(row=row, column=_NOTES_COLUMN, value=_NOTES_LABEL)
    notes_header.font = _HEADER_FONT
    notes_header.fill = _HEADER_FILL
    row += 1

    for paw in PAW_NODES:
        bad_by_stat = {stat: is_bad(paw, stat) for stat, _label, _fmt in stat_columns}
        bad = any(bad_by_stat.values())
        warn = (not bad) and is_warn(paw)
        name_cell = ws.cell(row=row, column=1, value=_PAW_LABELS[paw])
        if bad:
            name_cell.fill = _BAD_FILL
            name_cell.font = _BAD_FONT
        elif warn:
            name_cell.fill = _WARN_FILL
            name_cell.font = _WARN_FONT
        for col, (stat, _label, fmt) in enumerate(stat_columns, start=2):
            _write_cell(ws, row, col, get_value(stat, paw), fmt, bad=bad_by_stat[stat])
        note_cell = ws.cell(row=row, column=_NOTES_COLUMN, value=note_getter(paw) or None)
        if bad:
            note_cell.fill = _BAD_FILL
            note_cell.font = _BAD_FONT
        elif warn:
            note_cell.fill = _WARN_FILL
            note_cell.font = _WARN_FONT
        row += 1

    return row + 2  # blank spacer before the next block


def _format_sheet(ws):
    ws.column_dimensions["A"].width = 22
    for col in range(2, _N_COLUMNS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions[get_column_letter(_NOTES_COLUMN)].width = 40
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right")


def _safe_sheet_name(name: str, used: set) -> str:
    """Sanitize and de-duplicate an Excel sheet name (31-char limit, no ``[]:*?/\\``)."""
    cleaned = "".join(c if c not in _INVALID_SHEET_CHARS else "_" for c in str(name)).strip() or "rat"
    base = cleaned[:31]
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def write_group_report(
    trials: List[TrialMetrics],
    output_path: PathLike,
    manual_flags: Optional[Dict[str, Dict[int, Tuple[set, str]]]] = None,
) -> None:
    """Write one group's gait-metrics workbook: one tab per distinct ``rat_id``.

    Each tab stacks one titled block per trial (session) for that rat,
    followed, when a rat has more than one crossing, by a final "Average"
    block. A paw with no usable run (see :func:`_paw_has_no_usable_run`)
    has its whole row highlighted; a single column that failed on its own
    (see :func:`_stat_is_untrustworthy`) is highlighted by itself. The
    Average block's per-(paw, stat) means are NaN-aware.

    Args:
        manual_flags: Session name -> crossing number -> (flagged paw
            names, note), carried forward from
            :func:`alligaitor.validation.load_group_manual_flags` so a
            regenerated workbook keeps highlighting a paw a reviewer
            already flagged invalid by hand on that specific crossing.
            This only ever adds highlighting, never removes it.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    manual_flags = manual_flags or {}

    by_rat: Dict[str, List[TrialMetrics]] = {}
    for trial in trials:
        by_rat.setdefault(trial.rat_id, []).append(trial)

    wb = Workbook()
    wb.remove(wb.active)  # replaced by a real sheet per rat below (or the placeholder, if there are none)

    if not by_rat:
        ws = wb.create_sheet("Sheet1")
        ws.cell(row=1, column=1, value="No sessions in this group.")
    else:
        used_names: set = set()
        for rat_id, rat_trials in by_rat.items():
            ws = wb.create_sheet(_safe_sheet_name(rat_id, used_names))
            row = 1
            for trial in rat_trials:
                flagged_paws, note = manual_flags.get(trial.session_name, {}).get(
                    trial.crossing_index + 1, (set(), "")
                )
                row = _write_paw_block(
                    ws, row,
                    title=f"Session: {trial.crossing_label}",
                    crossing_time_s=trial.crossing_time_s,
                    average_speed_cm_s=trial.average_speed_mm_s / 10.0,
                    get_value=lambda stat, paw, t=trial: getattr(t, stat)[paw],
                    is_bad=lambda paw, stat, t=trial, fp=flagged_paws: (
                        _paw_has_no_usable_run(t, paw)
                        or paw in fp
                        or _stat_is_untrustworthy(t, paw, stat)
                    ),
                    is_warn=lambda paw, t=trial: (
                        t.bottom_fallback_fraction.get(paw, 0.0) > BOTTOM_FALLBACK_WARN_THRESHOLD
                    ),
                    note_getter=lambda paw, fp=flagged_paws, n=note, t=trial: (
                        n if paw in fp
                        else (
                            f"{t.bottom_fallback_fraction.get(paw, 0.0):.0%} of this run came from the "
                            "2D bottom-camera fallback"
                            if t.bottom_fallback_fraction.get(paw, 0.0) > BOTTOM_FALLBACK_WARN_THRESHOLD
                            else ""
                        )
                    ),
                )

            if len(rat_trials) > 1:
                avg_stat_columns = tuple((stat, label, "0.00") for stat, label, _fmt in _STAT_COLUMNS)
                sessions_here = {t.session_name for t in rat_trials}
                if len(sessions_here) == 1:
                    avg_title = f"Average of {len(rat_trials)} crossings"
                else:
                    avg_title = f"Average of {len(rat_trials)} crossings across {len(sessions_here)} sessions"
                row = _write_paw_block(
                    ws, row,
                    title=avg_title,
                    crossing_time_s=_nanmean([t.crossing_time_s for t in rat_trials]),
                    average_speed_cm_s=_nanmean([t.average_speed_mm_s for t in rat_trials]) / 10.0,
                    get_value=lambda stat, paw: _nanmean([getattr(t, stat)[paw] for t in rat_trials]),
                    is_bad=lambda paw, stat: all(
                        _paw_has_no_usable_run(t, paw)
                        or paw in manual_flags.get(t.session_name, {}).get(t.crossing_index + 1, (set(), ""))[0]
                        or _stat_is_untrustworthy(t, paw, stat)
                        for t in rat_trials
                    ),
                    is_warn=lambda paw: any(
                        t.bottom_fallback_fraction.get(paw, 0.0) > BOTTOM_FALLBACK_WARN_THRESHOLD
                        for t in rat_trials
                    ),
                    stat_columns=avg_stat_columns,
                )

            _format_sheet(ws)

    wb.save(output_path)


def _find_paw_row(ws, title: str, paw: str) -> Optional[int]:
    """Row index of `paw`'s row within the block whose title cell reads
    exactly `title` (see :func:`crossing_block_title`), or ``None`` if
    that block or paw row isn't present. Searches a bounded window below
    the title row rather than assuming :func:`_write_paw_block`'s exact
    row offsets.
    """
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value != title:
            continue
        for r in range(cell.row + 1, cell.row + 16):
            if ws.cell(row=r, column=1).value == _PAW_LABELS[paw]:
                return r
        return None
    return None


def annotate_manual_flag(
    xlsx_path: PathLike,
    rat_id: str,
    session_name: str,
    crossing_number: int,
    crossing_count: int,
    paw: str,
    auto_usable: bool,
    flagged: bool,
    note: str = "",
) -> bool:
    """Patch one paw's row, on one specific crossing, in an already-written
    group workbook to reflect a reviewer's manual flag -- without
    regenerating the whole report. A flag on one crossing never touches
    another crossing's block for the same paw.

    The row is highlighted (and `note` written to its Notes cell) when
    ``flagged`` or ``not auto_usable`` -- unflagging a paw the automatic
    detection already called unusable leaves it highlighted, since that
    finding wasn't created by this manual action and shouldn't be erased
    by it.

    Args:
        crossing_number: 1-based crossing index within the recording.
        crossing_count: How many crossings this recording has in total,
            needed to reconstruct the exact block title.

    Returns:
        ``True`` if the target row was found and patched, ``False`` if the
        workbook has no matching rat sheet / crossing block / paw row (the
        caller should warn rather than assume the flag took effect).
    """
    from openpyxl import load_workbook

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return False

    wb = load_workbook(xlsx_path)
    candidate = _safe_sheet_name(rat_id, set())
    ws = wb[candidate] if candidate in wb.sheetnames else next(
        (wb[name] for name in wb.sheetnames if name.startswith(candidate[: max(1, len(candidate) - 2)])), None
    )
    if ws is None:
        return False

    title = crossing_block_title(session_name, crossing_number, crossing_count)
    row = _find_paw_row(ws, title, paw)
    if row is None:
        return False

    bad = flagged or not auto_usable
    fill = _BAD_FILL if bad else _CLEAR_FILL
    font = _BAD_FONT if bad else _CLEAR_FONT
    for col in range(1, _N_STAT_COLUMNS + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
    notes_cell = ws.cell(row=row, column=_NOTES_COLUMN, value=(note if flagged else None))
    notes_cell.fill = fill
    notes_cell.font = font

    wb.save(xlsx_path)
    return True
